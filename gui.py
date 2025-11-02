#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Interfaz Gráfica para Sistema de Gestión de Pagos
"""

import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import os
from payment_manager import PaymentManager

# Intentar importar tkinterdnd2
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False


class PaymentGUI:
    """Interfaz gráfica para el sistema de gestión de pagos"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Gestión de Pagos")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)
        
        self.manager = PaymentManager()
        
        self.colors = {
            'bg_primary': '#ffffff',
            'bg_secondary': '#fafafa',
            'text_primary': '#424242',
            'text_secondary': '#757575',
            'border': '#e0e0e0'
        }
        
        self.setup_ui()
        
    def setup_ui(self):
        """Configura la interfaz principal"""
        self.root.configure(bg=self.colors['bg_primary'])
        
        self.setup_styles()
        self.setup_header()
        self.setup_payment_zone()
        self.setup_confirmation_zone()
        self.setup_monto_zone()
        self.setup_logs()
        self.setup_buttons()
        # Verificar estado inicial de zona de montos
        self.update_monto_zone_state()
        
    def setup_styles(self):
        """Configura los estilos de los widgets"""
        style = ttk.Style()
        style.theme_use('clam')
        
        style.configure('Title.TLabel',
                       font=('Segoe UI', 16, 'bold'),
                       background=self.colors['bg_primary'],
                       foreground=self.colors['text_primary'])
        
        style.configure('Zone.TLabel',
                       font=('Segoe UI', 12, 'bold'),
                       background=self.colors['bg_secondary'],
                       foreground=self.colors['text_primary'])
        
        style.configure('Info.TLabel',
                       font=('Segoe UI', 9),
                       background=self.colors['bg_primary'],
                       foreground=self.colors['text_secondary'])
        
        style.configure('Action.TButton',
                       font=('Segoe UI', 10),
                       padding=10)
        
        style.map('Action.TButton',
                 foreground=[('active', '#ffffff'),
                           ('pressed', '#ffffff')],
                 background=[('active', '#1565c0'),
                           ('pressed', '#0d47a1')])
        
    def setup_header(self):
        """Configura el encabezado"""
        header = tk.Frame(self.root, bg=self.colors['bg_primary'], pady=20)
        header.pack(fill=tk.X)
        
        title = ttk.Label(header, 
                         text="Sistema de Gestión de Pagos WhatsApp",
                         style='Title.TLabel')
        title.pack()
        
    def setup_payment_zone(self):
        """Configura la zona de carga de pagos (izquierda)"""
        container = tk.Frame(self.root, bg=self.colors['bg_primary'])
        container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        frame = tk.Frame(container, bg=self.colors['bg_secondary'], 
                        relief=tk.RAISED, bd=2)
        frame.pack(fill=tk.BOTH, expand=True)
        
        title = ttk.Label(frame, text="Subir Pagos", style='Zone.TLabel')
        title.pack(pady=15)
        
        zone = tk.Frame(frame, bg=self.colors['bg_secondary'], 
                       relief=tk.SUNKEN, bd=1, width=400, height=200)
        zone.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        zone_label = tk.Label(zone, 
                             text="Arrastra archivos .txt aquí\n\n"
                                 "o haz clic para seleccionar archivos",
                             bg=self.colors['bg_secondary'],
                             fg=self.colors['text_secondary'],
                             font=('Segoe UI', 10),
                             justify=tk.CENTER)
        zone_label.pack(expand=True)
        
        self.payment_zone = zone
        self.payment_label = zone_label
        
        self.payment_zone.bind("<Button-1>", lambda e: self.select_payment_files())
        self.payment_label.bind("<Button-1>", lambda e: self.select_payment_files())
        
        if DND_AVAILABLE:
            try:
                zone.drop_target_register(DND_FILES)
                zone.dnd_bind('<<Drop>>', self.on_drop_payment)
            except:
                pass
        
        info = ttk.Label(frame, 
                        text="Formatos soportados: .txt\n"
                            "Se extraerán los pagos del archivo",
                        style='Info.TLabel')
        info.pack(pady=10)
        
    def setup_confirmation_zone(self):
        """Configura la zona de carga de confirmaciones (derecha)"""
        container = tk.Frame(self.root, bg=self.colors['bg_primary'])
        container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        frame = tk.Frame(container, bg=self.colors['bg_secondary'], 
                        relief=tk.RAISED, bd=2)
        frame.pack(fill=tk.BOTH, expand=True)
        
        title = ttk.Label(frame, text="Subir Confirmaciones", style='Zone.TLabel')
        title.pack(pady=15)
        
        zone = tk.Frame(frame, bg=self.colors['bg_secondary'], 
                       relief=tk.SUNKEN, bd=1, width=400, height=200)
        zone.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        zone_label = tk.Label(zone, 
                             text="Arrastra archivos .txt aquí\n\n"
                                 "o haz clic para seleccionar archivos",
                             bg=self.colors['bg_secondary'],
                             fg=self.colors['text_secondary'],
                             font=('Segoe UI', 10),
                             justify=tk.CENTER)
        zone_label.pack(expand=True)
        
        self.confirmation_zone = zone
        self.confirmation_label = zone_label
        
        self.confirmation_zone.bind("<Button-1>", lambda e: self.select_confirmation_files())
        self.confirmation_label.bind("<Button-1>", lambda e: self.select_confirmation_files())
        
        if DND_AVAILABLE:
            try:
                zone.drop_target_register(DND_FILES)
                zone.dnd_bind('<<Drop>>', self.on_drop_confirmation)
            except:
                pass
        
        info = ttk.Label(frame, 
                        text="Formatos soportados: .txt\n"
                            "Se marcarán como confirmados",
                        style='Info.TLabel')
        info.pack(pady=10)
        
    def setup_monto_zone(self):
        """Configura la zona de carga de Excel de montos"""
        container = tk.Frame(self.root, bg=self.colors['bg_primary'])
        container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        frame = tk.Frame(container, bg=self.colors['bg_secondary'], 
                        relief=tk.RAISED, bd=2)
        frame.pack(fill=tk.BOTH, expand=True)
        
        title = ttk.Label(frame, text="Subir Excel de Montos", style='Zone.TLabel')
        title.pack(pady=15)
        
        zone = tk.Frame(frame, bg=self.colors['bg_secondary'], 
                       relief=tk.SUNKEN, bd=1, width=400, height=200)
        zone.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        zone_label = tk.Label(zone, 
                             text="Primero genera Pagos.xlsx\nprocesando pagos",
                             bg=self.colors['bg_secondary'],
                             fg=self.colors['text_secondary'],
                             font=('Segoe UI', 10),
                             justify=tk.CENTER)
        zone_label.pack(expand=True)
        
        self.monto_zone = zone
        self.monto_label = zone_label
        
        self.monto_zone.bind("<Button-1>", lambda e: self.select_monto_file())
        self.monto_label.bind("<Button-1>", lambda e: self.select_monto_file())
        
        if DND_AVAILABLE:
            try:
                zone.drop_target_register(DND_FILES)
                zone.dnd_bind('<<Drop>>', self.on_drop_monto)
            except:
                pass
        
        info = ttk.Label(frame, 
                        text="Formato soportado: .xlsx\n"
                            "Archivo de montos autorizados",
                        style='Info.TLabel')
        info.pack(pady=10)
        
    def check_pagos_excel_exists(self) -> bool:
        """Verifica si existe el archivo Excel de pagos"""
        return os.path.exists(self.manager.excel_path)
    
    def update_monto_zone_state(self):
        """Habilita o deshabilita la zona de montos según exista Pagos.xlsx"""
        if self.check_pagos_excel_exists():
            # Habilitar zona - cambiar texto y color
            self.monto_label.config(
                text="Arrastra archivo Excel aquí\n\n"
                     "o haz clic para seleccionar archivo",
                fg=self.colors['text_primary']
            )
            # Habilitar eventos de click
            self.monto_zone.bind("<Button-1>", lambda e: self.select_monto_file())
            self.monto_label.bind("<Button-1>", lambda e: self.select_monto_file())
        else:
            # Deshabilitar zona - cambiar texto y color
            self.monto_label.config(
                text="Primero genera Pagos.xlsx\nprocesando pagos",
                fg=self.colors['text_secondary']
            )
            # Deshabilitar eventos de click
            self.monto_zone.unbind("<Button-1>")
            self.monto_label.unbind("<Button-1>")
        
    def setup_logs(self):
        """Configura el área de logs"""
        container = tk.Frame(self.root, bg=self.colors['bg_primary'])
        container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        label = ttk.Label(container, text="Log de Actividad", 
                         style='Zone.TLabel')
        label.pack(anchor=tk.W, pady=(0, 5))
        
        self.log_text = scrolledtext.ScrolledText(
            container,
            wrap=tk.WORD,
            width=80,
            height=10,
            bg='#ffffff',
            fg=self.colors['text_primary'],
            font=('Consolas', 9),
            relief=tk.SUNKEN,
            bd=1
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        self.log("Sistema iniciado correctamente")
        
    def setup_buttons(self):
        """Configura los botones de acción"""
        container = tk.Frame(self.root, bg=self.colors['bg_primary'])
        container.pack(fill=tk.X, padx=10, pady=10)
        
        btn_frame = tk.Frame(container, bg=self.colors['bg_primary'])
        btn_frame.pack()
        
        btn_excel = ttk.Button(btn_frame, 
                              text="Ver Excel",
                              style='Action.TButton',
                              command=self.view_excel)
        btn_excel.pack(side=tk.LEFT, padx=5)
        
        btn_clear = ttk.Button(btn_frame,
                              text="Limpiar Registros",
                              style='Action.TButton',
                              command=self.clear_data)
        btn_clear.pack(side=tk.LEFT, padx=5)
        
        btn_exit = ttk.Button(btn_frame,
                             text="Salir",
                             style='Action.TButton',
                             command=self.exit_app)
        btn_exit.pack(side=tk.LEFT, padx=5)
        
    def select_payment_files(self):
        """Abre diálogo para seleccionar archivos de pagos"""
        files = filedialog.askopenfilenames(
            title="Seleccionar Archivos de Pagos",
            filetypes=[("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*")]
        )
        if files:
            self.process_payments(files)
            
    def select_confirmation_files(self):
        """Abre diálogo para seleccionar archivos de confirmaciones"""
        files = filedialog.askopenfilenames(
            title="Seleccionar Archivos de Confirmaciones",
            filetypes=[("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*")]
        )
        if files:
            self.process_confirmations(files)
    
    def select_monto_file(self):
        """Abre diálogo para seleccionar archivo Excel de montos"""
        if not self.check_pagos_excel_exists():
            messagebox.showwarning(
                "Excel No Existe",
                "Primero debes generar el archivo Pagos.xlsx procesando algunos pagos."
            )
            return
        
        file = filedialog.askopenfilename(
            title="Seleccionar Archivo Excel de Montos",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        if file:
            self.process_monto_file(file)
            
    def on_drop_payment(self, event):
        """Maneja el evento de arrastrar y soltar en zona de pagos"""
        files = self.root.tk.splitlist(event.data)
        valid_files = []
        for filepath in files:
            if filepath.lower().endswith('.txt'):
                valid_files.append(filepath)
        if valid_files:
            self.process_payments(valid_files)
            
    def on_drop_confirmation(self, event):
        """Maneja el evento de arrastrar y soltar en zona de confirmaciones"""
        files = self.root.tk.splitlist(event.data)
        valid_files = []
        for filepath in files:
            if filepath.lower().endswith('.txt'):
                valid_files.append(filepath)
        if valid_files:
            self.process_confirmations(valid_files)
    
    def on_drop_monto(self, event):
        """Maneja el evento de arrastrar y soltar en zona de montos"""
        if not self.check_pagos_excel_exists():
            messagebox.showwarning(
                "Excel No Existe",
                "Primero debes generar el archivo Pagos.xlsx procesando algunos pagos."
            )
            return
        
        files = self.root.tk.splitlist(event.data)
        valid_files = []
        for filepath in files:
            if filepath.lower().endswith('.xlsx'):
                valid_files.append(filepath)
        if valid_files:
            self.process_monto_file(valid_files[0])  # Solo procesar el primer archivo
            
    def process_payments(self, filepaths):
        """Procesa los archivos de pagos"""
        self.log(f"Procesando {len(filepaths)} archivo(s) de pagos...")
        
        all_entries = []
        total_errors = 0
        total_duplicates = 0
        
        for filepath in filepaths:
            self.log(f"Archivo: {os.path.basename(filepath)}")
            
            entries, errors, duplicates = self.manager.process_file(filepath)
            
            all_entries.extend(entries)
            total_errors += errors
            total_duplicates += duplicates
            
            self.log(f"  -> Entradas extraídas: {len(entries)}")
            self.log(f"  -> Errores: {errors}")
            self.log(f"  -> Duplicados: {duplicates}")
        
        if all_entries:
            self.log("Agregando entradas al Excel...")
            num_added = self.manager.add_to_excel(all_entries)
            self.log(f"Total de registros en Excel: {num_added}")
            
            # Actualizar estado de zona de montos después de crear/actualizar Excel
            self.update_monto_zone_state()
            
            messagebox.showinfo(
                "Pagos Procesados",
                f"Se procesaron {len(all_entries)} entradas\n"
                f"Total de registros en Excel: {num_added}"
            )
        else:
            self.log("No se encontraron pagos válidos en los archivos")
            messagebox.showwarning(
                "Sin Resultados",
                "No se encontraron pagos válidos en los archivos seleccionados"
            )
            
    def process_confirmations(self, filepaths):
        """Procesa los archivos de confirmaciones"""
        self.log(f"Procesando {len(filepaths)} archivo(s) de confirmaciones...")
        
        all_confirmed = []
        all_alerts = []
        
        for filepath in filepaths:
            self.log(f"Archivo: {os.path.basename(filepath)}")
            
            confirmed, alerts = self.manager.process_confirmations(filepath)
            
            all_confirmed.extend(confirmed)
            all_alerts.extend(alerts)
            
            self.log(f"  -> Confirmaciones: {len(confirmed)}")
            self.log(f"  -> Alertas: {len(alerts)}")
        
        if all_alerts:
            self.log("Alertas encontradas:")
            for alert in all_alerts:
                self.log(f"  - {alert}")
                
        if all_confirmed:
            messagebox.showinfo(
                "Confirmaciones Procesadas",
                f"Se confirmaron {len(all_confirmed)} pagos\n"
                f"Excel actualizado correctamente"
            )
        else:
            if all_alerts:
                messagebox.showwarning(
                    "Problemas al Confirmar",
                    "No se pudieron confirmar los pagos.\n"
                    "Revisa el log para más detalles."
                )
            else:
                messagebox.showinfo(
                    "Procesado",
                    "El archivo de confirmaciones fue procesado."
                )
    
    def process_monto_file(self, filepath):
        """Procesa el archivo Excel de montos y actualiza Pagos.xlsx"""
        if not self.check_pagos_excel_exists():
            messagebox.showwarning(
                "Excel No Existe",
                "Primero debes generar el archivo Pagos.xlsx procesando algunos pagos."
            )
            return
        
        self.log(f"Procesando archivo de montos: {os.path.basename(filepath)}")
        
        # Cargar archivo de montos
        success = self.manager.load_monto_file(filepath)
        
        if not success:
            self.log("Error al cargar archivo de montos")
            messagebox.showerror(
                "Error",
                "No se pudo cargar el archivo de montos.\nRevisa el log para más detalles."
            )
            return
        
        # Actualizar Excel existente con valores de Pago semanal
        try:
            import pandas as pd
            
            # Leer Excel actual con dtype=str para preservar ceros a la izquierda
            df_pagos = pd.read_excel(
                self.manager.excel_path, 
                sheet_name='Pagos', 
                engine='openpyxl',
                dtype={'ID': str, 'Ciclo': str, 'Depósito': str}
            )
            
            # Normalizar ID (ya es string, solo asegurar formato)
            if 'ID' in df_pagos.columns:
                df_pagos['ID'] = df_pagos['ID'].astype(str).str.replace('.0', '', regex=False).str.replace('nan', '').str.replace('None', '')
                df_pagos['ID'] = df_pagos['ID'].str.zfill(6)
            
            # Normalizar Depósito (ya es string, solo asegurar formato de 9 dígitos)
            if 'Depósito' in df_pagos.columns:
                df_pagos['Depósito'] = df_pagos['Depósito'].astype(str).str.replace('.0', '', regex=False).str.replace('nan', '').str.replace('None', '')
                # Asegurar formato completo de 9 dígitos
                def fix_deposito(val):
                    if pd.isna(val) or val == '' or val == 'nan' or val == 'None':
                        return None
                    val_str = str(val).strip()
                    # Si es numérico y tiene menos de 9 dígitos, rellenar
                    if val_str.replace('.', '').isdigit():
                        val_str = val_str.split('.')[0]
                        if len(val_str) < 9:
                            val_str = val_str.zfill(9)
                    return val_str
                df_pagos['Depósito'] = df_pagos['Depósito'].apply(fix_deposito)
            
            # Actualizar o agregar columna Pago semanal
            df_pagos['Pago semanal'] = df_pagos.apply(
                lambda row: self.manager.get_pago_semanal(
                    str(row.get('ID', '')).zfill(6),
                    str(row.get('Tipo', 'Ind')).strip()
                ), axis=1
            )
            
            # Contar cuántos registros fueron actualizados
            registros_actualizados = len(df_pagos)
            registros_encontrados = len(df_pagos[df_pagos['Pago semanal'] != 'No encontrado'])
            
            # Guardar Excel actualizado
            cols_orden = ['Tipo', 'ID', 'Grupo', 'Fecha', 'Hora', 'Pago', 'Ahorro', 'Total', 
                         'Número de Pago', 'Sucursal', 'Corte', 'Ciclo', 'Concepto', 'Depósito', 'Confirmado', 'Pago semanal']
            
            # Asegurar todas las columnas existan
            for col in cols_orden:
                if col not in df_pagos.columns:
                    if col == 'Pago semanal':
                        df_pagos[col] = 'No encontrado'
                    else:
                        df_pagos[col] = None
            
            # Asegurar que Depósito sea string para preservar ceros a la izquierda
            if 'Depósito' in df_pagos.columns:
                df_pagos['Depósito'] = df_pagos['Depósito'].astype(str)
            
            # Reordenar columnas
            df_pagos = df_pagos.reindex(columns=cols_orden)
            
            # Guardar
            with pd.ExcelWriter(self.manager.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_pagos.to_excel(writer, sheet_name='Pagos', index=False)
            
            # Configurar formato de Depósito como texto en Excel
            import openpyxl
            wb = openpyxl.load_workbook(self.manager.excel_path)
            if 'Pagos' in wb.sheetnames:
                ws = wb['Pagos']
                for cell in ws[1]:  # Primera fila (encabezados)
                    if cell.value == 'Depósito':
                        col_letter = cell.column_letter
                        # Formatear todas las celdas de la columna Depósito como texto
                        for row in range(2, ws.max_row + 1):
                            cell_ref = ws[f'{col_letter}{row}']
                            cell_ref.number_format = '@'  # @ = texto
                            # Asegurar que el valor se guarde como string
                            if cell_ref.value is not None:
                                cell_ref.value = str(cell_ref.value)
            wb.save(self.manager.excel_path)
            wb.close()
            
            self.log(f"Archivo de montos procesado: {registros_encontrados}/{registros_actualizados} registros encontraron pago semanal")
            
            messagebox.showinfo(
                "Archivo de Montos Procesado",
                f"Se actualizó el archivo Pagos.xlsx\n\n"
                f"Registros actualizados: {registros_actualizados}\n"
                f"Pagos semanales encontrados: {registros_encontrados}"
            )
            
        except Exception as e:
            self.log(f"Error actualizando Excel con montos: {e}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror(
                "Error",
                f"Error al actualizar el Excel:\n{e}"
            )
                
    def log(self, message):
        """Agrega un mensaje al log"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
        
    def clear_data(self):
        """Limpia todos los registros del sistema"""
        if not messagebox.askyesno(
            "Confirmar Limpieza",
            "¿Estás seguro de que deseas eliminar TODOS los registros?\n\n"
            "Esto eliminará:\n"
            "- Archivo Excel (Pagos.xlsx)\n"
            "- Configuración (config.json)\n"
            "- Logs (log.txt)\n\n"
            "Esta acción NO se puede deshacer."
        ):
            return
        
        self.log("Limpiando todos los registros...")
        success = self.manager.clear_all_data()
        
        if success:
            self.log("Todos los datos fueron limpiados exitosamente")
            messagebox.showinfo("Limpieza Completa", 
                              "Todos los registros fueron eliminados exitosamente")
        else:
            self.log("Advertencia: Algunos archivos no pudieron ser eliminados")
            messagebox.showwarning(
                "Limpieza Parcial",
                "Algunos archivos no pudieron ser eliminados.\n"
                "Asegúrate de cerrar el archivo Excel si está abierto."
            )
            
    def view_excel(self):
        """Abre el archivo Excel en el programa predeterminado"""
        excel_path = self.manager.excel_path
        
        if not os.path.exists(excel_path):
            messagebox.showwarning(
                "Archivo No Encontrado",
                f"No existe el archivo {excel_path}\n\n"
                "Procesa algunos pagos primero para crear el archivo."
            )
            return
        
        try:
            os.startfile(excel_path)
            self.log(f"Abriendo {excel_path}")
        except Exception as e:
            self.log(f"Error al abrir Excel: {e}")
            messagebox.showerror(
                "Error",
                f"No se pudo abrir el archivo Excel:\n{e}"
            )
            
    def exit_app(self):
        """Cierra la aplicación"""
        if messagebox.askyesno("Salir", "¿Deseas salir de la aplicación?"):
            self.log("Cerrando aplicación...")
            self.root.quit()
            self.root.destroy()


def main():
    """Función principal"""
    # Usar TkinterDnD si está disponible, sino tk.Tk
    if DND_AVAILABLE:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    
    app = PaymentGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()