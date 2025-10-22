import json
import os
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelManager:
    """Gestiona la generación y actualización de archivos Excel"""
    
    def __init__(self, config_path: str = "config.json"):
        """Inicializa el gestor de Excel"""
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        output_dir = config['rutas']['output']
        os.makedirs(output_dir, exist_ok=True)
        
        self.excel_path = os.path.join(output_dir, 'pagos.xlsx')
    
    def crear_excel_desde_cero(self, pagos: List[Dict]):
        """Crea un nuevo archivo Excel con todos los pagos"""
        wb = Workbook()
        
        # Hoja principal de datos
        ws_datos = wb.active
        ws_datos.title = "Pagos"
        
        # Encabezados mejorados
        headers = [
            'ID Grupo', 'Grupo', 'Fecha', 'Hora', 
            'Pago', 'Ahorro', 'Total', 'Corte Horario'
        ]
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws_datos.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Ordenar pagos por grupo y fecha
        pagos_ordenados = sorted(pagos, key=lambda x: (x['id_grupo'], x['fecha_mensaje']))
        
        # Escribir datos
        for row_num, pago in enumerate(pagos_ordenados, 2):
            fecha_mensaje = datetime.fromisoformat(pago['fecha_mensaje']) if isinstance(pago['fecha_mensaje'], str) else pago['fecha_mensaje']
            
            total = pago['pago'] + pago['ahorro']
            
            ws_datos.cell(row=row_num, column=1, value=pago['id_grupo'])
            ws_datos.cell(row=row_num, column=2, value=pago['grupo'])
            ws_datos.cell(row=row_num, column=3, value=fecha_mensaje.strftime('%d/%m/%Y'))
            ws_datos.cell(row=row_num, column=4, value=fecha_mensaje.strftime('%H:%M:%S'))
            ws_datos.cell(row=row_num, column=5, value=pago['pago'])
            ws_datos.cell(row=row_num, column=6, value=pago['ahorro'])
            ws_datos.cell(row=row_num, column=7, value=total)
            ws_datos.cell(row=row_num, column=8, value=pago['corte_horario'])
        
        # Formato de moneda para columnas de Pago, Ahorro y Total
        for row in range(2, len(pagos) + 2):
            ws_datos.cell(row=row, column=5).number_format = '$#,##0.00'
            ws_datos.cell(row=row, column=6).number_format = '$#,##0.00'
            ws_datos.cell(row=row, column=7).number_format = '$#,##0.00'
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 12,  # ID Grupo
            'B': 25,  # Grupo
            'C': 13,  # Fecha
            'D': 11,  # Hora
            'E': 14,  # Pago
            'F': 14,  # Ahorro
            'G': 14,  # Total
            'H': 18   # Corte Horario
        }
        
        for col, width in column_widths.items():
            ws_datos.column_dimensions[col].width = width
        
        # Habilitar filtros
        ws_datos.auto_filter.ref = f"A1:H{len(pagos) + 1}"
        
        # Hoja de resumen
        self._crear_hoja_resumen(wb, pagos)
        
        # Guardar
        wb.save(self.excel_path)
    
    def _crear_hoja_resumen(self, wb: Workbook, pagos: List[Dict]):
        """Crea una hoja con resumen por grupos y estadísticas"""
        ws_resumen = wb.create_sheet("Resumen por Grupos")
        
        # Título
        ws_resumen['A1'] = "RESUMEN POR GRUPOS"
        ws_resumen['A1'].font = Font(bold=True, size=16)
        ws_resumen['A1'].alignment = Alignment(horizontal="center")
        ws_resumen.merge_cells('A1:F1')
        
        # Fecha de generación
        ws_resumen['A3'] = "Fecha de generación:"
        ws_resumen['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        ws_resumen['A3'].font = Font(bold=True)
        
        # Total de registros
        ws_resumen['A4'] = "Total de registros:"
        ws_resumen['B4'] = len(pagos)
        ws_resumen['A4'].font = Font(bold=True)
        
        # Cálculos totales
        total_pago = sum(p['pago'] for p in pagos)
        total_ahorro = sum(p['ahorro'] for p in pagos)
        total_general = total_pago + total_ahorro
        
        ws_resumen['A5'] = "Total Pagos:"
        ws_resumen['B5'] = total_pago
        ws_resumen['B5'].number_format = '$#,##0.00'
        ws_resumen['A5'].font = Font(bold=True)
        
        ws_resumen['A6'] = "Total Ahorros:"
        ws_resumen['B6'] = total_ahorro
        ws_resumen['B6'].number_format = '$#,##0.00'
        ws_resumen['A6'].font = Font(bold=True)
        
        ws_resumen['A7'] = "Total General:"
        ws_resumen['B7'] = total_general
        ws_resumen['B7'].number_format = '$#,##0.00'
        ws_resumen['A7'].font = Font(bold=True, color="FF0000")
        
        # Resumen por corte
        ws_resumen['A8'] = "RESUMEN POR CORTE HORARIO"
        ws_resumen['A8'].font = Font(bold=True, size=14)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws_resumen['A9'] = "Corte"
        ws_resumen['B9'] = "Cantidad"
        ws_resumen['C9'] = "Total Pago"
        ws_resumen['D9'] = "Total Ahorro"
        
        for cell in ['A9', 'B9', 'C9', 'D9']:
            ws_resumen[cell].fill = header_fill
            ws_resumen[cell].font = header_font
        
        # Agrupar por corte
        cortes = {}
        for pago in pagos:
            corte = pago['corte_horario']
            if corte not in cortes:
                cortes[corte] = {'cantidad': 0, 'pago': 0, 'ahorro': 0}
            cortes[corte]['cantidad'] += 1
            cortes[corte]['pago'] += pago['pago']
            cortes[corte]['ahorro'] += pago['ahorro']
        
        row = 10
        for corte, data in cortes.items():
            ws_resumen.cell(row=row, column=1, value=corte)
            ws_resumen.cell(row=row, column=2, value=data['cantidad'])
            ws_resumen.cell(row=row, column=3, value=data['pago'])
            ws_resumen.cell(row=row, column=4, value=data['ahorro'])
            ws_resumen.cell(row=row, column=3).number_format = '$#,##0.00'
            ws_resumen.cell(row=row, column=4).number_format = '$#,##0.00'
            row += 1
        
        # Resumen por grupos
        row += 2
        ws_resumen.cell(row=row, column=1, value="RESUMEN POR GRUPOS")
        ws_resumen.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        
        # Headers para grupos
        ws_resumen.cell(row=row, column=1, value="ID Grupo")
        ws_resumen.cell(row=row, column=2, value="Grupo")
        ws_resumen.cell(row=row, column=3, value="Cantidad")
        ws_resumen.cell(row=row, column=4, value="Total Pago")
        ws_resumen.cell(row=row, column=5, value="Total Ahorro")
        ws_resumen.cell(row=row, column=6, value="Total General")
        
        for col in range(1, 7):
            ws_resumen.cell(row=row, column=col).fill = header_fill
            ws_resumen.cell(row=row, column=col).font = header_font
        
        # Agrupar por grupo
        grupos = {}
        for pago in pagos:
            id_grupo = pago['id_grupo']
            if id_grupo not in grupos:
                grupos[id_grupo] = {
                    'nombre': pago['grupo'],
                    'cantidad': 0,
                    'pago': 0,
                    'ahorro': 0
                }
            grupos[id_grupo]['cantidad'] += 1
            grupos[id_grupo]['pago'] += pago['pago']
            grupos[id_grupo]['ahorro'] += pago['ahorro']
        
        row += 1
        for id_grupo in sorted(grupos.keys()):
            data = grupos[id_grupo]
            total_grupo = data['pago'] + data['ahorro']
            
            ws_resumen.cell(row=row, column=1, value=id_grupo)
            ws_resumen.cell(row=row, column=2, value=data['nombre'])
            ws_resumen.cell(row=row, column=3, value=data['cantidad'])
            ws_resumen.cell(row=row, column=4, value=data['pago'])
            ws_resumen.cell(row=row, column=5, value=data['ahorro'])
            ws_resumen.cell(row=row, column=6, value=total_grupo)
            
            ws_resumen.cell(row=row, column=4).number_format = '$#,##0.00'
            ws_resumen.cell(row=row, column=5).number_format = '$#,##0.00'
            ws_resumen.cell(row=row, column=6).number_format = '$#,##0.00'
            row += 1
        
        # Ajustar anchos
        ws_resumen.column_dimensions['A'].width = 12
        ws_resumen.column_dimensions['B'].width = 25
        ws_resumen.column_dimensions['C'].width = 12
        ws_resumen.column_dimensions['D'].width = 16
        ws_resumen.column_dimensions['E'].width = 16
        ws_resumen.column_dimensions['F'].width = 16
    
    def generar_excel(self, pagos: List[Dict]):
        """Genera o actualiza el archivo Excel con los pagos"""
        if not pagos:
            print("No hay pagos para generar en Excel")
            return
        
        # Siempre crear desde cero para tener datos actualizados
        self.crear_excel_desde_cero(pagos)
        print(f"Excel generado exitosamente: {self.excel_path}")
        print(f"Total de registros: {len(pagos)}")


if __name__ == "__main__":
    # Prueba básica
    excel_manager = ExcelManager()
    print(f"Excel Manager inicializado")
    print(f"Ruta de salida: {excel_manager.excel_path}")

