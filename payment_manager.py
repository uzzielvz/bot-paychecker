#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Gestión de Pagos desde WhatsApp
Extrae registros de pagos de chats .txt y los gestiona en Excel
"""

import re
import os
import sys
import json
import logging
import time
from datetime import datetime
from typing import List, Dict, Tuple, Optional, Generator
import unicodedata

try:
    import pandas as pd
except ImportError:
    print("Error: pandas no está instalado. Ejecuta: pip install pandas")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl no está instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)


class PaymentManager:
    """Gestiona el parsing, normalización y almacenamiento de pagos"""
    
    def __init__(self, excel_path="Pagos.xlsx"):
        self.excel_path = excel_path
        self.config_path = "config.json"
        self.setup_logging()
        self.load_config()
        # Diccionarios para lookup de pago semanal desde archivo de montos
        self.monto_grupos = {}  # {cod_grupo_solidario: valor_AC}
        self.monto_individuales = {}  # {codigo_acreditado: valor_AC}
        
    def load_config(self):
        """Carga configuración desde config.json, si no existe el json, se crea uno por defecto"""
        self.config = {
            "horarios": {
                "matutino": "< 13:00",
                "vespertino": ">= 13:00",
                "archivo_procesado": None,
                "corte_actual": None
            },
            "mapeo_id_grupos": {}
        }
        
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
            except Exception as e:
                logging.error(f"Error cargando config: {e}")
    
    def save_config(self):
        """Guarda configuración a config.json"""
        try:
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logging.error(f"Error guardando config: {e}")
    
    def get_group_info_from_config(self, payment_id: str) -> Tuple[Optional[str], Optional[str]]:
        """
        Obtiene nombre y sucursal normalizados desde config.json
        Retorna: (nombre_normalizado, sucursal)
        """
        if payment_id in self.config.get("mapeo_id_grupos", {}):
            grupo_info = self.config["mapeo_id_grupos"][payment_id]
            return grupo_info.get("nombre"), grupo_info.get("sucursal")
        return None, None
    
    def load_monto_file(self, monto_filepath: str) -> bool:
        """
        Carga archivo Excel de montos autorizados y crea diccionarios de lookup.
        Para grupales: Columna C "Cod. grupo solidario" -> Columna AC "Parcialidad + Parcialidad comisión"
        Para individuales: Columna A "Codigo acreditado" -> Columna AC "Parcialidad + Parcialidad comisión"
        """
        try:
            if not os.path.exists(monto_filepath):
                logging.error(f"Archivo de montos no encontrado: {monto_filepath}")
                return False
            
            # Leer Excel
            df = pd.read_excel(monto_filepath, engine='openpyxl')
            
            # Limpiar diccionarios anteriores
            self.monto_grupos = {}
            self.monto_individuales = {}
            
            # Obtener nombres de columnas (pueden tener caracteres especiales)
            # Columna A: índice 0, Columna C: índice 2, Columna AC: índice 28
            col_a_idx = 0  # "Codigo acreditado"
            col_c_idx = 2  # "Cod. grupo solidario"
            col_ac_idx = 28  # "Parcialidad + Parcialidad comisión"
            
            if len(df.columns) <= col_ac_idx:
                logging.error(f"El archivo Excel no tiene suficientes columnas. Se esperaba al menos columna AC (índice 28)")
                return False
            
            # Procesar cada fila
            for idx, row in df.iterrows():
                # Obtener valor de columna AC (Parcialidad + Parcialidad comisión)
                valor_ac = row.iloc[col_ac_idx]
                
                # Si el valor es NaN, saltar esta fila
                if pd.isna(valor_ac):
                    continue
                
                # Convertir a string para almacenar (puede ser numérico)
                valor_ac_str = str(valor_ac).strip() if not pd.isna(valor_ac) else None
                if not valor_ac_str or valor_ac_str == 'nan':
                    continue
                
                # Procesar grupales: Columna C "Cod. grupo solidario"
                cod_grupo = row.iloc[col_c_idx]
                if pd.notna(cod_grupo):
                    # Convertir a string y normalizar (puede ser float, convertir a int primero si aplica)
                    try:
                        if isinstance(cod_grupo, float):
                            # Si es float entero, convertir a int y luego a string
                            if cod_grupo.is_integer():
                                cod_grupo_str = str(int(cod_grupo)).zfill(6)
                            else:
                                cod_grupo_str = str(int(cod_grupo)).zfill(6)
                        else:
                            cod_grupo_str = str(cod_grupo).strip().zfill(6)
                        
                        # Solo agregar si no existe (primera coincidencia, ya que todas tienen mismo valor)
                        if cod_grupo_str not in self.monto_grupos:
                            self.monto_grupos[cod_grupo_str] = valor_ac_str
                    except Exception as e:
                        logging.warning(f"Error procesando código grupo en fila {idx}: {e}")
                        continue
                
                # Procesar individuales: Columna A "Codigo acreditado"
                cod_acreditado = row.iloc[col_a_idx]
                if pd.notna(cod_acreditado):
                    # Convertir a string y normalizar
                    try:
                        if isinstance(cod_acreditado, float):
                            if cod_acreditado.is_integer():
                                cod_acreditado_str = str(int(cod_acreditado)).zfill(6)
                            else:
                                cod_acreditado_str = str(int(cod_acreditado)).zfill(6)
                        else:
                            cod_acreditado_str = str(cod_acreditado).strip().zfill(6)
                        
                        # Solo agregar si no existe
                        if cod_acreditado_str not in self.monto_individuales:
                            self.monto_individuales[cod_acreditado_str] = valor_ac_str
                    except Exception as e:
                        logging.warning(f"Error procesando código acreditado en fila {idx}: {e}")
                        continue
            
            logging.info(f"Archivo de montos cargado: {len(self.monto_grupos)} grupos, {len(self.monto_individuales)} individuales")
            return True
            
        except Exception as e:
            logging.error(f"Error cargando archivo de montos: {e}")
            import traceback
            logging.error(traceback.format_exc())
            return False
        
    def setup_logging(self):
        """Configura el logging a archivo"""
        logging.basicConfig(
            filename='log.txt',
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        
    def get_pago_semanal(self, payment_id: str, tipo: str) -> str:
        """
        Obtiene el pago semanal desde los diccionarios de montos cargados.
        tipo debe ser 'Gpo' o 'Ind'
        """
        # Normalizar payment_id a string con 6 dígitos
        payment_id_normalized = str(payment_id).strip().zfill(6)
        
        if tipo == 'Gpo':
            # Buscar en diccionario de grupos
            if self.monto_grupos and payment_id_normalized in self.monto_grupos:
                return str(self.monto_grupos[payment_id_normalized])
        elif tipo == 'Ind':
            # Buscar en diccionario de individuales
            if self.monto_individuales and payment_id_normalized in self.monto_individuales:
                return str(self.monto_individuales[payment_id_normalized])
        
        return "No encontrado"
    
    def extract_full_name(self, content: str) -> Optional[str]:
        """
        Extrae el nombre completo del grupo o cliente sin truncar.
        Captura TODO el texto entre 'Grupo:'/'Nombre Grupo:'/'Cliente:' y 'ID'.
        Usa patrones greedy para capturar múltiples palabras completas.
        Soporta formatos con asteriscos markdown.
        """
        # Patrón para Grupo o Nombre Grupo (soporta asteriscos opcionales antes)
        # Captura TODO hasta encontrar "ID" seguido de número (puede tener asteriscos antes de ID)
        grupo_pattern = r'(?:\*+\s*)?\*?\s*(?:Nombre\s+)?(?:Grupo|GRUPO)[:\s]+(.+?)\s+(?:\*+\s*)?\*?\s*ID[:\s]+\d+'
        grupo_match = re.search(grupo_pattern, content, re.IGNORECASE | re.DOTALL)
        
        if grupo_match:
            nombre = grupo_match.group(1).strip()
            # Limpiar asteriscos markdown, saltos de línea y espacios múltiples
            nombre = re.sub(r'\*+\s*', '', nombre)
            nombre = re.sub(r'\s+', ' ', nombre).strip()
            if nombre:
                return nombre.upper()
        
        # Patrón para Cliente (soporta asteriscos opcionales)
        cliente_pattern = r'(?:\*+\s*)?\*?\s*Cliente[:\s]+(.+?)\s+(?:\*+\s*)?\*?\s*ID[:\s]+\d+'
        cliente_match = re.search(cliente_pattern, content, re.IGNORECASE | re.DOTALL)
        
        if cliente_match:
            nombre = cliente_match.group(1).strip()
            # Limpiar asteriscos markdown, saltos de línea y espacios múltiples
            nombre = re.sub(r'\*+\s*', '', nombre)
            nombre = re.sub(r'\s+', ' ', nombre).strip()
            if nombre:
                return nombre.upper()
        
        return None
    
    def normalize_sucursal(self, text: str) -> str:
        """Quita acentos de las sucursales"""
        if not text or text.strip() == '':
            return "Sin especificar"
        text = text.strip()
        nfd = unicodedata.normalize('NFD', text)
        ascii_text = nfd.encode('ascii', 'ignore').decode('ascii')
        return ascii_text
    
    def normalize_number(self, text: str) -> float:
        """Normaliza números quitando $, comas y convirtiendo a float"""
        if not text:
            return 0.0
        cleaned = re.sub(r'[\$,\s]', '', str(text))
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    
    def get_current_corte(self) -> str:
        """
        Determina el corte horario actual basado en la hora del sistema
        Matutino < 13:00, Vespertino >= 13:00
        """
        hora_actual = datetime.now().hour
        corte = "Matutino" if hora_actual < 13 else "Vespertino"
        
        # Guardar en config
        self.config["horarios"]["corte_actual"] = corte
        self.save_config()
        
        return corte
    
    def extract_all_payments_from_lines(self, lines: List[str], filename: str, corte: str = None) -> List[Dict]:
        """Extrae todos los pagos de las líneas del archivo"""
        entries = []
        # Soporta formato con/sin p.m./a.m. y horas con 1 o 2 dígitos
        msg_pattern = r'\[(\d{2}/\d{2}/\d{2}), (\d{1,2}:\d{2}:\d{2})\s*(?:a\.m\.|p\.m\.)?\] ([^:]+): (.+)'
        
        i = 0
        current_fecha = None
        current_hora = None
        
        while i < len(lines):
            line = lines[i]
            match = re.match(msg_pattern, line)
            
            if match:
                current_fecha = match.group(1)
                current_hora = match.group(2)
                content = match.group(4)
                
                # Acumular líneas siguientes hasta el siguiente mensaje
                following_lines = []
                j = i + 1
                while j < len(lines) and not re.match(msg_pattern, lines[j]):
                    following_lines.append(lines[j].strip())
                    j += 1
                
                # Combinar contenido
                full_content = content + '\n' + '\n'.join(following_lines)
                
                # Extraer grupos de este mensaje
                extracted = self.extract_payments_from_content(
                    full_content, current_fecha, current_hora, filename, corte
                )
                entries.extend(extracted)
                
                i = j
            else:
                i += 1
        
        return entries
    
    def extract_payments_from_content(self, content: str, fecha: str, hora: str, filename: str, corte: str = None) -> List[Dict]:
        """Extrae uno o más pagos del contenido de un mensaje"""
        entries = []
        
        # Ignorar mensajes del sistema (solo si el contenido COMPLETO es un mensaje del sistema)
        # No ignorar si contiene información de pago válida
        if content.strip() in ['Creaste el grupo', 'Los mensajes y las llamadas están cifrados de extremo a extremo. Solo las personas en este chat pueden leerlos, escucharlos o compartirlos.', '']:
            return entries
        # Ignorar solo si el contenido empieza con estos textos y no tiene datos de pago
        if (content.startswith('Creaste el grupo') or content.startswith('Los mensajes y las llamadas están cifrados')) and not re.search(r'(?:Grupo|Cliente|ID\s*\d|Pago)', content, re.IGNORECASE):
            return entries
        
        # Detectar tipo: Individual (Cliente o formato ID+NOMBRE) o Grupal (Grupo)
        es_individual_cliente = bool(re.search(r'\bCliente\b', content, re.IGNORECASE))
        
        # Detectar individuales sin "Cliente": formato "001395 ROMANO PALMA EDITH YADIRA"
        # Regex busca ID al inicio o después de timestamp, seguido de nombre en mayúsculas
        ind_match_sin_cliente = re.search(r'^\s*0*(\d{6})\s+([A-ZÁÉÍÓÚÑ\s]+?)(?:\s*\(|$)', content.strip(), re.MULTILINE)
        es_individual_sin_cliente = ind_match_sin_cliente is not None
        
        es_individual = es_individual_cliente or es_individual_sin_cliente
        es_grupal = bool(re.search(r'\bGrupo\b|\bGRUPO\b', content, re.IGNORECASE))
        
        # Si no hay ni Cliente, ni formato ID+NOMBRE, ni Grupo, no procesar
        if not es_individual and not es_grupal:
            return entries
        
        # Si es individual (con o sin Cliente), usar extract_single_payment
        if es_individual:
            single_entry = self.extract_single_payment(content, fecha, hora, filename, corte)
            if single_entry:
                entries.append(single_entry)
            return entries
        
        # Buscar todos los grupos en el contenido (solo para grupales)
        # Usar extract_full_name para capturar nombres completos sin truncar
        # Buscar primero dónde están los grupos para procesarlos individualmente
        grupo_nombre_pattern = r'(?:Grupo|Nombre\s+Grupo|GRUPO)\s*:?\s*'
        grupo_positions = list(re.finditer(grupo_nombre_pattern, content, re.IGNORECASE))
        
        if not grupo_positions:
            # Intentar extraer un solo grupo
            single_entry = self.extract_single_payment(content, fecha, hora, filename, corte)
            if single_entry:
                entries.append(single_entry)
            return entries
        
        # Para cada posición de grupo encontrada, extraer el nombre completo
        for grupo_pos_match in grupo_positions:
            try:
                grupo_start = grupo_pos_match.start()
                # Extraer el contenido desde este grupo hasta el siguiente o fin
                siguiente_grupo_match = re.search(r'(?:Grupo|Nombre\s+Grupo)\s*:?\s*', content[grupo_start+1:], re.IGNORECASE)
                if siguiente_grupo_match:
                    grupo_content = content[grupo_start:siguiente_grupo_match.start()+grupo_start+1]
                else:
                    grupo_content = content[grupo_start:]
                
                # Extraer nombre completo usando extract_full_name
                grupo = self.extract_full_name(grupo_content)
                if not grupo:
                    # Fallback al patrón anterior si extract_full_name falla
                    grupo_match = re.search(r'(?:\*+\s*)?\*?\s*(?:Nombre\s+)?(?:Grupo|GRUPO)\s*:?\s*([A-Za-zÀ-ÿ\s]+?)(?:\s|$|\*|:)', grupo_content, re.IGNORECASE)
                    if grupo_match:
                        grupo = grupo_match.group(1).strip().upper()
                    else:
                        continue
                
                # Buscar ID después del nombre del grupo (puede estar en línea separada)
                # Buscar en las siguientes líneas después del grupo, hasta el siguiente grupo o fin de contenido
                content_after_grupo = content[grupo_start:]
                # Buscar el siguiente grupo para delimitar la búsqueda
                siguiente_grupo_match = re.search(r'(?:\*+\s*)?\*?\s*(?:Nombre\s+)?(?:Grupo|GRUPO)\s*:?\s*', content_after_grupo[1:], re.IGNORECASE)
                if siguiente_grupo_match:
                    search_window = content_after_grupo[:siguiente_grupo_match.start()+1]
                else:
                    search_window = content_after_grupo[:1000]  # Buscar hasta 1000 caracteres
                
                # Buscar ID con varios formatos en la ventana de búsqueda
                # Soporta: * **ID:**, **ID:**, ID Grupo, ID:, ID
                # El formato * **ID:** tiene: asterisco, espacio, dos asteriscos, ID, dos puntos, más asteriscos opcionales
                id_match = re.search(r'\*\s+\*\*ID\*\*\s*:?\s*0*(\d{1,6})|\*\s+\*\*ID\s*:?\s*\*+\s*0*(\d{1,6})|\*\*\s*ID\s*\*\*\s*:?\s*0*(\d{1,6})|\*\*ID\*\*\s*:?\s*0*(\d{1,6})|\*+\s*\*?\s*ID\s*:?\s*\*?\s*0*(\d{1,6})|ID\s+(?:Grupo\s+)?0*(\d{1,6})|ID\s*:?\s*0*(\d{1,6})', search_window, re.IGNORECASE)
                if not id_match:
                    continue
                
                payment_id = (id_match.group(1) or id_match.group(2) or id_match.group(3) or id_match.group(4) or id_match.group(5) or id_match.group(6) or id_match.group(7)).zfill(6)
                
                # Extraer datos después del ID encontrado (relativo a la posición del grupo)
                id_relative_pos = id_match.end()
                start_pos = grupo_start + id_relative_pos
                
                # Buscar Pago (soporta asteriscos markdown: * **Pago:**, **Pago:**, Pago:)
                # El formato * **Pago:** tiene asteriscos separados por espacio
                pago_match = re.search(r'\*\s+\*\*\s*Pago\s*\*?\s*:?\s*\*?\s*\$?\s*([\d,\.]+)|\*\*Pago\*\*\s*:?\s*\$?\s*([\d,\.]+)|\*+\s*\*?\s*Pago\s*:?\s*\*?\s*\$?\s*([\d,\.]+)', content[start_pos:], re.IGNORECASE)
                if not pago_match:
                    # Intentar sin asteriscos
                    pago_match = re.search(r'Pago\s*:?\s*\$?\s*([\d,\.]+)', content[start_pos:], re.IGNORECASE)
                if not pago_match:
                    continue
                pago = self.normalize_number(pago_match.group(1) or pago_match.group(2) or pago_match.group(3) or pago_match.group(1))
                
                # Buscar Ahorro (soporta asteriscos markdown: * **Ahorro: $X, **Ahorro:**, Ahorro: $X)
                # El formato * **Ahorro: $X tiene asteriscos separados por espacio
                ahorro_match = re.search(r'\*\s+\*\*\s*Ahorro\s*\*?\s*:?\s*\$\s*([\d,\.]+)|\*\*Ahorro\*\*\s*:?\s*\$\s*([\d,\.]+)|\*+\s*\*?\s*Ahorro\s*:?\s*\$\s*([\d,\.]+)', content[start_pos:], re.IGNORECASE)
                if not ahorro_match:
                    # Intentar con asteriscos pero sin el $ explícito
                    ahorro_match = re.search(r'\*+\s*\*?\s*Ahorro\s*:?\s*\$?\s*([\d,\.]+)', content[start_pos:], re.IGNORECASE)
                if not ahorro_match:
                    # Intentar sin asteriscos
                    ahorro_match = re.search(r'Ahorro\s*:?\s*\$?\s*([\d,\.]+)', content[start_pos:], re.IGNORECASE)
                ahorro = self.normalize_number(ahorro_match.group(1) or ahorro_match.group(2) or ahorro_match.group(3) or ahorro_match.group(1)) if ahorro_match else 0.0
                
                # Buscar Sucursal (soporta asteriscos markdown)
                sucursal_match = re.search(r'\*+\s*\*?\s*Sucursal\s*:?\s*\*?\s*([A-Za-zÀ-ÿ\s]+?)(?=\s*(?:N[úu]mero|$))', content[start_pos:], re.IGNORECASE)
                if not sucursal_match:
                    # Intentar sin asteriscos
                    sucursal_match = re.search(r'Sucursal\s*:?\s*([A-Za-zÀ-ÿ\s]+?)(?=\s*(?:N[úu]mero|$))', content[start_pos:], re.IGNORECASE)
                sucursal = sucursal_match.group(1).strip() if sucursal_match else None
                
                # Buscar Número de pago (soporta "Pago semana X" y "Número de pago: X" con asteriscos)
                num_match = re.search(r'\*+\s*\*?\s*(?:Número de pago|N[úu]mero de pago|N pago|N Pago)\s*:?\s*\*?\s*(\d+)', content[start_pos:], re.IGNORECASE)
                if not num_match:
                    # Intentar sin asteriscos
                    num_match = re.search(r'(?:Pago\s+semana|Número de pago|N[úu]mero de pago|N pago|N Pago)\s*:?\s*(\d+)', content[start_pos:], re.IGNORECASE)
                if not num_match:
                    # Intentar formato corto "Pago X"
                    num_match = re.search(r'Pago\s+(\d+)(?:\s|$)', content[start_pos:], re.IGNORECASE)
                num_pago = int(num_match.group(1)) if num_match else None
                # Si no hay número de pago, usar "Pendiente"
                if num_pago is None:
                    num_pago = "Pendiente"
                
                # Buscar Ciclo (OBLIGATORIO, solo acepta 1 o 2) - soporta asteriscos markdown
                # Buscar primero en todo el content (puede estar fuera del bloque del grupo)
                ciclo_match = re.search(r'Ciclo\s*:?\s*0?(\d+)', content, re.IGNORECASE)
                if not ciclo_match:
                    ciclo_match = re.search(r'\*\*Ciclo\*\*\s*0?(\d+)', content, re.IGNORECASE)
                if not ciclo_match:
                    ciclo_match = re.search(r'\*+\s*\*?\s*Ciclo\s*:?\s*0?(\d+)', content, re.IGNORECASE)
                if not ciclo_match:
                    logging.warning(f"Ciclo no encontrado para ID {payment_id}")
                    continue
                
                ciclo_num = int(ciclo_match.group(1))
                if ciclo_num not in [1, 2]:
                    logging.warning(f"Ciclo inválido {ciclo_num} para ID {payment_id}")
                    continue
                
                ciclo_formato = f"{ciclo_num:02d}"
                
                # Calcular Concepto Depósito: tipo_code(1) + ID(6) + Ciclo(2)
                tipo_code = '0'  # Es grupal (cambio de '2' a '0')
                id_str = payment_id.zfill(6)
                ciclo_str = ciclo_formato.zfill(2)
                deposito = tipo_code + id_str + ciclo_str
                
                # Intentar obtener info normalizada del config
                nombre_config, sucursal_config = self.get_group_info_from_config(payment_id)
                
                # Usar corte horario actual o determinar desde hora
                if corte is None:
                    hora_int = int(hora.split(':')[0]) if ':' in hora else 12
                    corte = "Matutino" if hora_int < 12 else "Vespertino"
                
                # Calcular Total
                total_calculado = round(pago + ahorro, 2)
                
                # Buscar Total en el contenido para validación
                total_match = re.search(r'Total\s*:?\s*\$?\s*([\d,\.]+)', content, re.IGNORECASE)
                if total_match:
                    total_dado = self.normalize_number(total_match.group(1))
                    # Validar que Total = Pago + Ahorro (tolerancia 0.01)
                    if abs(total_dado - total_calculado) > 0.01:
                        logging.warning(f"Discrepancia en Total para ID {payment_id}: "
                                      f"Calculado={total_calculado}, Dado={total_dado}, "
                                      f"Diferencia={abs(total_dado - total_calculado)}")
                
                entry = {
                    'Tipo': 'Gpo',  # Es grupal
                    'ID': payment_id,
                    'Grupo': nombre_config if nombre_config else grupo.upper(),
                    'Fecha': fecha,
                    'Hora': hora,
                    'Pago': round(pago, 2),
                    'Ahorro': round(ahorro, 2),
                    'Total': total_calculado,
                    'Número de Pago': num_pago,
                    'Sucursal': sucursal_config if sucursal_config else (self.normalize_sucursal(sucursal) if sucursal else "Pendiente"),
                    'Corte': corte,
                    'Ciclo': ciclo_formato,  # Formato "01" o "02"
                    'Concepto': "Pendiente de imagen",  # Default para grupales
                    'Depósito': deposito,  # Calculado: tipo(1) + ID(6) + Ciclo(2)
                    'Confirmado': 'No',
                    'Pago semanal': self.get_pago_semanal(payment_id, 'Gpo'),
                    'Archivo': filename
                }
                
                entries.append(entry)
            except Exception as e:
                logging.error(f"Error parseando entrada: {e}")
                continue
        
        return entries
    
    def extract_single_payment(self, content: str, fecha: str, hora: str, filename: str, corte: str = None) -> Optional[Dict]:
        """Extrae un solo pago del contenido (Individual o Grupal)"""
        # Detectar tipo: Individual (Cliente o formato ID+NOMBRE) o Grupal (Grupo)
        es_individual_cliente = bool(re.search(r'\bCliente\b', content, re.IGNORECASE))
        
        # Detectar individuales sin "Cliente": formato "001395 ROMANO PALMA EDITH YADIRA"
        ind_match_sin_cliente = re.search(r'^\s*0*(\d{6})\s+([A-ZÁÉÍÓÚÑ\s]+?)(?:\s*\(|$)', content.strip(), re.MULTILINE)
        es_individual_sin_cliente = ind_match_sin_cliente is not None
        
        es_individual = es_individual_cliente or es_individual_sin_cliente
        es_grupal = bool(re.search(r'\bGrupo\b|\bGRUPO\b', content, re.IGNORECASE))
        
        if not es_individual and not es_grupal:
            return None
        
        # Buscar ID según el formato
        payment_id = None
        nombre_ind_sin_cliente = None
        concepto_ind_sin_cliente = None
        
        if es_individual_sin_cliente:
            # Formato: "001395 ROMANO PALMA EDITH YADIRA" o "001395 ROMANO PALMA EDITH YADIRA (NOTA)"
            payment_id = ind_match_sin_cliente.group(1).zfill(6)
            nombre_ind_sin_cliente = ind_match_sin_cliente.group(2).strip().upper()
            
            # Extraer Concepto si hay paréntesis
            concepto_match = re.search(r'\(([^)]+)\)', content)
            if concepto_match:
                concepto_ind_sin_cliente = concepto_match.group(1).strip()
            
        # Si no se encontró con formato nuevo, buscar formato tradicional (soporta "ID Grupo" y "ID:")
        if not payment_id:
            id_match = re.search(r'ID\s+(?:Grupo\s+)?0*(\d{1,6})|ID\s*:?\s*0*(\d{1,6})', content, re.IGNORECASE)
            if not id_match:
                return None
            payment_id = (id_match.group(1) or id_match.group(2)).zfill(6)
        
        # Buscar Pago (OPCIONAL para individuales sin Cliente, requerido para otros, soporta asteriscos)
        pago_match = re.search(r'\*+\s*\*?\s*Pago\s*:?\s*\*?\s*\$?\s*([\d,\.]+)', content, re.IGNORECASE)
        if not pago_match:
            pago_match = re.search(r'Pago\s*:?\s*\$?\s*([\d,\.]+)', content, re.IGNORECASE)
        if pago_match:
            pago = self.normalize_number(pago_match.group(1))
        else:
            # Pago no encontrado - solo permitido para individuales sin Cliente
            if es_individual_sin_cliente:
                pago = 0.0
                logging.info(f"ID {payment_id}: Pago no encontrado → 0.0 (pendiente de imagen)")
            else:
                return None  # Para otros formatos, Pago es obligatorio
        
        # Buscar Sucursal
        sucursal_match = re.search(r'Sucursal\s*:?\s*([A-Za-zÀ-ÿ\s]+?)(?=\s*(?:N[úu]mero|$))', content)
        sucursal = sucursal_match.group(1).strip() if sucursal_match else None
        # Default inteligente: si no hay sucursal, usar "Pendiente"
        if not sucursal:
            sucursal = "Pendiente"
        
        # Buscar Ciclo (OBLIGATORIO, solo acepta 1 o 2, default "01" si falta, soporta asteriscos)
        ciclo_match = re.search(r'\*+\s*\*?\s*Ciclo\s*\*?\s*:?\s*0?(\d+)', content, re.IGNORECASE)
        if not ciclo_match:
            ciclo_match = re.search(r'Ciclo\s*:?\s*0?(\d+)', content, re.IGNORECASE)
        if not ciclo_match:
            # Default: usar "01" si no se encuentra (solo para individuales sin Cliente)
            if es_individual_sin_cliente:
                ciclo_num = 1
                ciclo_formato = "01"
                logging.info(f"ID {payment_id}: Ciclo no encontrado → usando default '01'")
            else:
                logging.warning(f"Ciclo inválido o faltante para ID {payment_id}: No encontrado")
                return None
        else:
            ciclo_num = int(ciclo_match.group(1))
            if ciclo_num not in [1, 2]:
                # Default: usar "01" si es inválido (solo para individuales sin Cliente)
                if es_individual_sin_cliente:
                    ciclo_num = 1
                    ciclo_formato = "01"
                    logging.warning(f"ID {payment_id}: Ciclo inválido {ciclo_match.group(1)} → usando default '01'")
                else:
                    logging.warning(f"Ciclo inválido o faltante para ID {payment_id}: {ciclo_num}")
                    return None
            else:
                ciclo_formato = f"{ciclo_num:02d}"  # "01" o "02"
        
        # Calcular Concepto Depósito: tipo_code(1) + ID(6) + Ciclo(2)
        # Se determinará el tipo_code según si es Ind o Gpo
        # id_str ya está en formato de 6 dígitos (payment_id)
        id_str = payment_id.zfill(6)
        ciclo_str = ciclo_formato.zfill(2)
        
        # Intentar obtener info normalizada del config
        nombre_config, sucursal_config = self.get_group_info_from_config(payment_id)
        
        # Usar corte horario actual o determinar desde hora
        if corte is None:
            hora_int = int(hora.split(':')[0]) if ':' in hora else 12
            corte = "Matutino" if hora_int < 12 else "Vespertino"
        
        # Procesar según tipo
        if es_individual:
            # INDIVIDUAL: Determinar nombre según formato
            if es_individual_sin_cliente:
                # Formato: "001395 ROMANO PALMA EDITH YADIRA"
                grupo = nombre_config if nombre_config else nombre_ind_sin_cliente
                concepto = concepto_ind_sin_cliente if concepto_ind_sin_cliente else "Pendiente de imagen"
            else:
                # Formato tradicional: "Cliente NOMBRE ID..."
                # Usar extract_full_name para capturar nombre completo sin truncar
                cliente_nombre = self.extract_full_name(content)
                if not cliente_nombre:
                    # Fallback al patrón anterior si extract_full_name falla
                    cliente_match = re.search(r'Cliente\s+([A-Za-zÀ-ÿ\s]+?)(?=\s+ID)', content, re.IGNORECASE)
                    if not cliente_match:
                        return None
                    cliente_nombre = cliente_match.group(1).strip().upper()
                grupo = nombre_config if nombre_config else cliente_nombre
                concepto = "Pendiente de imagen"  # Default para formato tradicional
            
            # Reglas para Individual
            ahorro = 0.0  # Siempre 0.0 para Ind
            total_calculado = round(pago, 2)  # Total = Pago
            num_pago = None  # Sin número de pago
            tipo = 'Ind'
            tipo_code = '1'  # Individual
        else:
            # GRUPAL: Buscar Grupo (soporta asteriscos markdown y "ID Grupo")
            # Usar extract_full_name para capturar nombre completo sin truncar
            grupo = self.extract_full_name(content)
            if not grupo:
                # Fallback al patrón anterior si extract_full_name falla
                grupo_match = re.search(r'(?:\*+\s*)?\*?\s*(?:Nombre\s+)?(?:Grupo|GRUPO)\s*:?\s*([A-Za-zÀ-ÿ\s]+?)(?=\s*(?:ID|ID\s+Grupo|\d{6}))', content, re.IGNORECASE)
                if not grupo_match:
                    return None
                grupo = grupo_match.group(1).strip().upper()
            
            # Buscar Ahorro (solo para grupales, soporta asteriscos markdown)
            ahorro_match = re.search(r'\*+\s*\*?\s*Ahorro\s*:?\s*\$?\s*([\d,\.]+)', content, re.IGNORECASE)
            if not ahorro_match:
                ahorro_match = re.search(r'Ahorro\s*:?\s*\$?\s*([\d,\.]+)', content, re.IGNORECASE)
            ahorro = self.normalize_number(ahorro_match.group(1)) if ahorro_match else 0.0
            
            # Buscar Número de pago (solo para grupales, soporta "Pago semana X")
            num_match = re.search(r'(?:Pago\s+semana|Número de pago|N[úu]mero de pago|N pago|N Pago)\s*:?\s*(\d+)', content, re.IGNORECASE)
            if not num_match:
                # Intentar formato corto "Pago X"
                num_match = re.search(r'Pago\s+(\d+)(?:\s|$)', content, re.IGNORECASE)
            num_pago = int(num_match.group(1)) if num_match else None
            # Si no hay número de pago y es grupal, usar "Pendiente"
            if es_grupal and num_pago is None:
                num_pago = "Pendiente"
            
            # Calcular Total para grupal
            total_calculado = round(pago + ahorro, 2)
            tipo = 'Gpo'
            tipo_code = '0'  # Grupal (cambio de '2' a '0')
            concepto = "Pendiente de imagen"  # Default para grupales
        
        # Calcular Concepto Depósito: tipo_code(1) + ID(6) + Ciclo(2)
        deposito = tipo_code + id_str + ciclo_str
        
        # Buscar Total en el contenido para validación (solo grupal puede tenerlo)
        if es_grupal:
            total_match = re.search(r'Total\s*:?\s*\$?\s*([\d,\.]+)', content, re.IGNORECASE)
            if total_match:
                total_dado = self.normalize_number(total_match.group(1))
                # Validar que Total = Pago + Ahorro (tolerancia 0.01)
                if abs(total_dado - total_calculado) > 0.01:
                    logging.warning(f"Discrepancia en Total para ID {payment_id}: "
                                  f"Calculado={total_calculado}, Dado={total_dado}, "
                                  f"Diferencia={abs(total_dado - total_calculado)}")
        
        return {
            'Tipo': tipo,
            'ID': payment_id,
            'Grupo': nombre_config if nombre_config else grupo.upper(),
            'Fecha': fecha,
            'Hora': hora,
            'Pago': round(pago, 2),
            'Ahorro': round(ahorro, 2),
            'Total': total_calculado,
            'Número de Pago': num_pago,
            'Sucursal': sucursal_config if sucursal_config else (self.normalize_sucursal(sucursal) if sucursal else "Pendiente"),
            'Corte': corte,
            'Ciclo': ciclo_formato,  # Formato "01" o "02"
            'Concepto': concepto,  # Extraído del texto o "Pendiente de imagen"
            'Depósito': deposito,  # Calculado: tipo_code(1) + ID(6) + Ciclo(2)
            'Confirmado': 'No',
            'Pago semanal': self.get_pago_semanal(payment_id, tipo),
            'Archivo': filename
        }
    
    def get_last_timestamp(self) -> Optional[str]:
        """Obtiene el último timestamp procesado desde la hoja Meta"""
        try:
            if not os.path.exists(self.excel_path):
                return None
            
            df_meta = pd.read_excel(self.excel_path, sheet_name='Meta', engine='openpyxl')
            if df_meta.empty or 'ultimo_timestamp' not in df_meta.columns:
                return None
            
            return df_meta.iloc[0]['ultimo_timestamp']
        except Exception as e:
            logging.error(f"Error leyendo último timestamp: {e}")
            return None
    
    def save_timestamp(self, timestamp: str):
        """Guarda el último timestamp procesado en la hoja Meta"""
        try:
            wb = openpyxl.load_workbook(self.excel_path) if os.path.exists(self.excel_path) else openpyxl.Workbook()
            
            # Eliminar hoja Meta si existe
            if 'Meta' in wb.sheetnames:
                wb.remove(wb['Meta'])
            
            ws_meta = wb.create_sheet('Meta', 0)
            ws_meta.append(['ultimo_timestamp'])
            ws_meta.append([timestamp])
            ws_meta.sheet_state = 'hidden'
            
            wb.save(self.excel_path)
            wb.close()
        except Exception as e:
            logging.error(f"Error guardando timestamp: {e}")
    
    def extract_last_timestamp_from_file(self, filepath: str) -> Optional[str]:
        """Extrae el timestamp del último mensaje en el archivo"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # Soporta formato con/sin p.m./a.m. y horas con 1 o 2 dígitos
            msg_pattern = r'\[(\d{2}/\d{2}/\d{2}), (\d{1,2}:\d{2}:\d{2})\s*(?:a\.m\.|p\.m\.)?\]'
            for line in reversed(lines):
                match = re.search(msg_pattern, line)
                if match:
                    fecha = match.group(1)
                    hora = match.group(2)
                    # Limpiar "p.m./a.m." si estaba presente en la captura
                    hora = hora.split()[0] if ' ' in hora else hora
                    dd, mm, yy = fecha.split('/')
                    timestamp = f"{yy}/{mm}/{dd} {hora}"
                    return timestamp
            return None
        except Exception as e:
            logging.error(f"Error extrayendo timestamp de {filepath}: {e}")
            return None
    
    def process_file(self, filepath: str) -> Tuple[List[Dict], int, int]:
        """Procesa un archivo .txt y extrae pagos"""
        entries = []
        errors = 0
        duplicates = 0
        
        # Verificar si el archivo ya fue procesado
        last_ts = self.extract_last_timestamp_from_file(filepath)
        if last_ts:
            stored_ts = self.get_last_timestamp()
            if stored_ts and last_ts <= stored_ts:
                logging.info(f"Archivo {filepath} ya procesado")
                return [], 0, 1
        
        # Obtener corte horario actual
        corte_actual = self.get_current_corte()
        
        try:
            # Usar generador línea por línea en lugar de readlines
            lines = []
            with open(filepath, 'r', encoding='utf-8') as f:
                for line in f:
                    lines.append(line.rstrip('\n\r'))
            
            filename = os.path.basename(filepath)
            entries = self.extract_all_payments_from_lines(lines, filename, corte_actual)
            
            # Eliminar duplicados usando ID + Grupo + Pago + Ahorro + timestamp
            seen = set()
            unique_entries = []
            for entry in entries:
                key = f"{entry['ID']}_{entry['Grupo']}_{entry['Pago']}_{entry['Ahorro']}_{entry['Fecha']} {entry['Hora']}"
                if key not in seen:
                    seen.add(key)
                    unique_entries.append(entry)
                else:
                    duplicates += 1
            
            entries = unique_entries
            
            # Guardar timestamp si se procesó exitosamente
            if entries and last_ts:
                self.save_timestamp(last_ts)
                logging.info(f"Procesados {len(entries)} pagos de {filepath}")
                
                # Guardar corte de procesamiento en config
                self.config["horarios"]["archivo_procesado"] = corte_actual
                self.save_config()
            
        except Exception as e:
            logging.error(f"Error procesando {filepath}: {e}")
            errors += 1
        
        return entries, errors, duplicates
    
    def add_to_excel(self, entries: List[Dict]) -> int:
        """Agrega entradas al Excel"""
        if not entries:
            logging.info("No hay entradas para agregar")
            return 0
        
        try:
            logging.info(f"Creando DataFrame con {len(entries)} entradas")
            df_new = pd.DataFrame(entries)
            
            # Orden EXACTO de columnas con 'Tipo' como primera columna, 'Concepto' después de 'Ciclo', 'Depósito' antes de 'Confirmado', 'Pago semanal' al final
            cols_orden = ['Tipo', 'ID', 'Grupo', 'Fecha', 'Hora', 'Pago', 'Ahorro', 'Total', 
                         'Número de Pago', 'Sucursal', 'Corte', 'Ciclo', 'Concepto', 'Depósito', 'Confirmado', 'Pago semanal']
            
            # Eliminar 'Archivo' que no debe ir al Excel
            if 'Archivo' in df_new.columns:
                df_new = df_new.drop(columns=['Archivo'])
            
            # Validar y filtrar entradas sin Ciclo válido
            valid_entries = []
            for idx, row in df_new.iterrows():
                ciclo_val = row.get('Ciclo')
                if pd.isna(ciclo_val) or str(ciclo_val).strip() == '':
                    logging.warning(f"Entrada descartada: Ciclo faltante para ID {row.get('ID', 'N/A')}")
                    continue
                
                # Validar que Ciclo sea "01" o "02"
                ciclo_str = str(ciclo_val).strip()
                if ciclo_str not in ['01', '02', '1', '2']:
                    logging.warning(f"Entrada descartada: Ciclo inválido '{ciclo_str}' para ID {row.get('ID', 'N/A')}")
                    continue
                
                # Normalizar a formato "01" o "02"
                if ciclo_str == '1':
                    ciclo_str = '01'
                elif ciclo_str == '2':
                    ciclo_str = '02'
                
                row['Ciclo'] = ciclo_str
                valid_entries.append(row)
            
            if not valid_entries:
                logging.warning("No hay entradas válidas después de validar Ciclo")
                return 0
            
            df_new = pd.DataFrame(valid_entries)
            
            # Asegurar columna 'Concepto' para entradas nuevas
            if 'Concepto' not in df_new.columns:
                df_new['Concepto'] = 'Pendiente de imagen'
            
            # Procesar columna 'Pago semanal' para nuevos registros
            if 'Pago semanal' not in df_new.columns:
                # Agregar columna aplicando lookup si hay diccionarios cargados
                df_new['Pago semanal'] = df_new.apply(
                    lambda row: self.get_pago_semanal(
                        str(row.get('ID', '')).zfill(6),
                        str(row.get('Tipo', 'Ind')).strip()
                    ), axis=1
                )
            else:
                # Si existe pero tiene valores vacíos, rellenar con lookup
                mask = (df_new['Pago semanal'].isna()) | (df_new['Pago semanal'] == '') | (df_new['Pago semanal'] == 'No encontrado')
                if mask.any():
                    df_new.loc[mask, 'Pago semanal'] = df_new.loc[mask].apply(
                        lambda row: self.get_pago_semanal(
                            str(row.get('ID', '')).zfill(6),
                            str(row.get('Tipo', 'Ind')).strip()
                        ), axis=1
                    )
            
            # Calcular columna 'Depósito' para entradas nuevas si no existe
            if 'Depósito' not in df_new.columns:
                df_new['Depósito'] = df_new.apply(
                    lambda row: (
                        ('1' if str(row.get('Tipo', 'Ind')).strip() == 'Ind' else '0') +
                        str(row.get('ID', '')).zfill(6) +
                        str(row.get('Ciclo', '01')).zfill(2)
                    ), axis=1
                )
            
            # Asegurar que todas las columnas existan (rellenar con valores por defecto si faltan)
            for col in cols_orden:
                if col not in df_new.columns:
                    if col == 'Tipo':
                        # Si no hay Tipo, inferir de otros campos
                        df_new[col] = df_new.apply(
                            lambda row: 'Gpo' if pd.notna(row.get('Ahorro', 0)) and float(row.get('Ahorro', 0)) > 0 
                                       else 'Ind', axis=1
                        )
                    elif col == 'Ciclo':
                        # Ciclo es obligatorio, no debería faltar pero por seguridad
                        logging.warning("Columna Ciclo faltante en datos - esto no debería pasar")
                        continue
                    elif col == 'Concepto':
                        # Default: "Pendiente de imagen"
                        df_new[col] = 'Pendiente de imagen'
                    elif col == 'Pago semanal':
                        # Calcular Pago semanal si falta
                        df_new[col] = df_new.apply(
                            lambda row: self.get_pago_semanal(
                                str(row.get('ID', '')).zfill(6),
                                str(row.get('Tipo', 'Ind')).strip()
                            ), axis=1
                        )
                    elif col == 'Depósito':
                        # Calcular Depósito si falta
                        df_new[col] = df_new.apply(
                            lambda row: (
                                ('1' if str(row.get('Tipo', 'Ind')).strip() == 'Ind' else '0') +
                                str(row.get('ID', '')).zfill(6) +
                                str(row.get('Ciclo', '01')).zfill(2)
                            ), axis=1
                        )
                    else:
                        df_new[col] = None
            
            # Reordenar columnas al orden exacto especificado
            df_new = df_new.reindex(columns=cols_orden)
            
            # Convertir ID a string y asegurar formato de 6 dígitos con ceros a la izquierda
            if 'ID' in df_new.columns:
                df_new['ID'] = df_new['ID'].astype(str).str.zfill(6)
            
            # Verificar si existe archivo con hoja Pagos
            df_existing = None
            if os.path.exists(self.excel_path):
                try:
                    # Leer con dtype=str para columnas críticas para preservar ceros a la izquierda
                    df_existing = pd.read_excel(
                        self.excel_path, 
                        sheet_name='Pagos', 
                        engine='openpyxl',
                        dtype={'ID': str, 'Ciclo': str, 'Depósito': str}
                    )
                    
                    # Normalizar ID (ya es string por dtype, solo limpiar y formatear)
                    if 'ID' in df_existing.columns:
                        df_existing['ID'] = df_existing['ID'].astype(str).str.replace('.0', '', regex=False).str.replace('nan', '').str.replace('None', '')
                        df_existing['ID'] = df_existing['ID'].str.zfill(6)
                    
                    # Normalizar Depósito (ya es string por dtype, solo asegurar formato completo)
                    if 'Depósito' in df_existing.columns:
                        df_existing['Depósito'] = df_existing['Depósito'].astype(str).str.replace('.0', '', regex=False).str.replace('nan', '').str.replace('None', '')
                        # Asegurar formato completo de 9 dígitos (tipo(1) + ID(6) + Ciclo(2))
                        def fix_deposito_format(val):
                            if pd.isna(val) or val == '' or val == 'nan' or val == 'None':
                                return None
                            val_str = str(val).strip()
                            # Si es numérico y tiene menos de 9 dígitos, rellenar con ceros
                            if val_str.replace('.', '').isdigit():
                                # Quitar punto decimal si existe
                                val_str = val_str.split('.')[0]
                                # Rellenar a 9 dígitos si es necesario
                                if len(val_str) < 9:
                                    val_str = val_str.zfill(9)
                            return val_str
                        
                        df_existing['Depósito'] = df_existing['Depósito'].apply(fix_deposito_format)
                    
                    # Si Excel existente no tiene 'Tipo', agregarlo y rellenar
                    if 'Tipo' not in df_existing.columns:
                        # Inferir Tipo de campos existentes
                        df_existing['Tipo'] = df_existing.apply(
                            lambda row: 'Gpo' if pd.notna(row.get('Ahorro', 0)) and float(row.get('Ahorro', 0)) > 0 
                                       else 'Ind', axis=1
                        )
                    
                    # Si Excel existente no tiene 'Ciclo', agregarlo con valor por defecto "01"
                    if 'Ciclo' not in df_existing.columns:
                        df_existing['Ciclo'] = '01'
                        logging.info("Columna 'Ciclo' agregada a Excel existente con valor por defecto '01'")
                    
                    # Validar y filtrar entradas existentes sin Ciclo válido
                    valid_existing = []
                    for idx, row in df_existing.iterrows():
                        ciclo_val = row.get('Ciclo')
                        if pd.isna(ciclo_val) or str(ciclo_val).strip() == '':
                            # Si no tiene Ciclo, asignar "01" por defecto
                            row['Ciclo'] = '01'
                        else:
                            ciclo_str = str(ciclo_val).strip()
                            # Normalizar formato
                            if ciclo_str == '1':
                                ciclo_str = '01'
                            elif ciclo_str == '2':
                                ciclo_str = '02'
                            elif ciclo_str not in ['01', '02']:
                                # Ciclo inválido, asignar "01" por defecto
                                logging.warning(f"Ciclo inválido '{ciclo_str}' en Excel para ID {row.get('ID', 'N/A')}, asignando '01'")
                                ciclo_str = '01'
                            row['Ciclo'] = ciclo_str
                        valid_existing.append(row)
                    
                    df_existing = pd.DataFrame(valid_existing)
                    
                    # Si Excel existente no tiene 'Concepto', agregarlo con valor por defecto
                    if 'Concepto' not in df_existing.columns:
                        df_existing['Concepto'] = 'Pendiente de imagen'
                        logging.info("Columna 'Concepto' agregada a Excel existente con valor por defecto 'Pendiente de imagen'")
                    
                    # Calcular columna 'Depósito' para Excel existente (siempre recalcular)
                    df_existing['Depósito'] = df_existing.apply(
                        lambda row: (
                            ('1' if str(row.get('Tipo', 'Ind')).strip() == 'Ind' else '0') +
                            str(row.get('ID', '')).zfill(6) +
                            str(row.get('Ciclo', '01')).zfill(2)
                        ), axis=1
                    )
                    logging.info("Columna 'Depósito' recalculada para Excel existente")
                    
                    # Procesar columna 'Pago semanal' para Excel existente
                    if 'Pago semanal' not in df_existing.columns:
                        # Agregar columna aplicando lookup si hay diccionarios cargados
                        df_existing['Pago semanal'] = df_existing.apply(
                            lambda row: self.get_pago_semanal(
                                str(row.get('ID', '')).zfill(6),
                                str(row.get('Tipo', 'Ind')).strip()
                            ), axis=1
                        )
                        logging.info("Columna 'Pago semanal' agregada a Excel existente")
                    else:
                        # Actualizar valores faltantes o "No encontrado" si hay nuevos datos cargados
                        mask = (df_existing['Pago semanal'].isna()) | (df_existing['Pago semanal'] == '') | (df_existing['Pago semanal'] == 'No encontrado')
                        if mask.any():
                            df_existing.loc[mask, 'Pago semanal'] = df_existing.loc[mask].apply(
                                lambda row: self.get_pago_semanal(
                                    str(row.get('ID', '')).zfill(6),
                                    str(row.get('Tipo', 'Ind')).strip()
                                ), axis=1
                            )
                            logging.info(f"Columna 'Pago semanal' actualizada para {mask.sum()} registros existentes")
                    
                    # Asegurar todas las columnas del orden especificado
                    for col in cols_orden:
                        if col not in df_existing.columns:
                            if col == 'Ciclo':
                                df_existing[col] = '01'  # Valor por defecto
                            elif col == 'Concepto':
                                df_existing[col] = 'Pendiente de imagen'
                            elif col == 'Depósito':
                                # Calcular Depósito si falta
                                df_existing[col] = df_existing.apply(
                                    lambda row: (
                                        ('1' if str(row.get('Tipo', 'Ind')).strip() == 'Ind' else '0') +
                                        str(row.get('ID', '')).zfill(6) +
                                        str(row.get('Ciclo', '01')).zfill(2)
                                    ), axis=1
                                )
                            else:
                                df_existing[col] = None
                    
                    # Reordenar columnas existentes al orden exacto
                    df_existing = df_existing.reindex(columns=cols_orden)
                except Exception as e:
                    logging.warning(f"Error leyendo Excel existente: {e}")
                    df_existing = None
            
            if df_existing is not None and not df_existing.empty:
                df_final = pd.concat([df_existing, df_new]).drop_duplicates(
                    subset=['ID', 'Grupo', 'Pago', 'Ahorro'], keep='first'
                ).reset_index(drop=True)
            else:
                df_final = df_new
            
            logging.info(f"Guardando {len(df_final)} registros en {self.excel_path}")
            
            # Asegurar que Depósito sea string para preservar ceros a la izquierda
            if 'Depósito' in df_final.columns:
                df_final['Depósito'] = df_final['Depósito'].astype(str)
            
            # Crear ExcelWriter y agregar ambas hojas
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='Pagos', index=False)
                # Crear hoja Meta vacía
                df_meta = pd.DataFrame({'ultimo_timestamp': ['']})
                df_meta.to_excel(writer, sheet_name='Meta', index=False)
            
            # Configurar formato de Excel y ocultar Meta con retries
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    wb = openpyxl.load_workbook(self.excel_path)
                    
                    # Configurar columnas ID, Ciclo y Depósito como texto para preservar formato (ceros a la izquierda)
                    if 'Pagos' in wb.sheetnames:
                        ws = wb['Pagos']
                        for cell in ws[1]:  # Primera fila (encabezados)
                            if cell.value == 'ID':
                                col_letter = cell.column_letter
                                # Formatear todas las celdas de la columna ID como texto
                                for row in range(2, ws.max_row + 1):
                                    ws[f'{col_letter}{row}'].number_format = '@'  # @ = texto
                            elif cell.value == 'Ciclo':
                                col_letter = cell.column_letter
                                # Formatear todas las celdas de la columna Ciclo como texto
                                for row in range(2, ws.max_row + 1):
                                    ws[f'{col_letter}{row}'].number_format = '@'  # @ = texto
                            elif cell.value == 'Depósito':
                                col_letter = cell.column_letter
                                # Formatear todas las celdas de la columna Depósito como texto
                                for row in range(2, ws.max_row + 1):
                                    cell_ref = ws[f'{col_letter}{row}']
                                    cell_ref.number_format = '@'  # @ = texto
                                    # Asegurar que el valor se guarde como string (preserva ceros a la izquierda)
                                    if cell_ref.value is not None:
                                        # Convertir a string, preservando formato completo con ceros
                                        dep_value = str(cell_ref.value)
                                        # Si el valor es numérico y empieza con 0, preservarlo
                                        if dep_value.isdigit() and len(dep_value) == 9:
                                            # Ya tiene formato correcto (9 dígitos: tipo(1) + ID(6) + Ciclo(2))
                                            cell_ref.value = dep_value
                                        else:
                                            # Normalizar a string asegurando formato completo
                                            cell_ref.value = str(cell_ref.value).zfill(9) if len(str(cell_ref.value)) < 9 else str(cell_ref.value)
                    
                    # Ocultar hoja Meta
                    if 'Meta' in wb.sheetnames and len(wb.sheetnames) > 1:
                        meta_idx = wb.sheetnames.index('Meta')
                        wb.worksheets[meta_idx].sheet_state = 'hidden'
                    
                    wb.save(self.excel_path)
                    wb.close()
                    break
                except PermissionError as pe:
                    if attempt < max_retries - 1:
                        logging.warning(f"Intento {attempt + 1} de {max_retries}: Permiso denegado. Esperando...")
                        time.sleep(1)
                    else:
                        logging.error(f"NO se pudo guardar Excel tras {max_retries} intentos. Cierra el archivo en Excel.")
                        raise
                except Exception as meta_error:
                    logging.warning(f"No se pudo configurar formato del Excel: {meta_error}")
                    break
            
            logging.info(f"Guardado exitoso: {len(df_final)} registros")
            return len(df_final)
        except Exception as e:
            import traceback
            logging.error(f"Error agregando a Excel: {e}")
            logging.error(traceback.format_exc())
            return 0
    
    def process_confirmations(self, filepath: str) -> Tuple[List[Dict], List[str]]:
        """
        Procesa archivo de confirmaciones y actualiza registros en Excel
        Retorna: (lista de confirmaciones procesadas, lista de alertas de no encontrados)
        """
        alerts = []
        confirmed_entries = []
        
        # Procesar archivo de confirmaciones directamente (sin filtro de timestamp)
        # Obtener corte horario actual para las confirmaciones
        corte_actual = self.get_current_corte()
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = []
                for line in f:
                    lines.append(line.rstrip('\n\r'))
            
            filename = os.path.basename(filepath)
            entries = self.extract_all_payments_from_lines(lines, filename, corte_actual)
            
            # Eliminar duplicados
            seen = set()
            unique_entries = []
            for entry in entries:
                key = f"{entry['ID']}_{entry['Grupo']}_{entry['Pago']}_{entry['Ahorro']}"
                if key not in seen:
                    seen.add(key)
                    unique_entries.append(entry)
            
            entries = unique_entries
        except Exception as e:
            logging.error(f"Error leyendo confirmaciones: {e}")
            alerts.append(f"Error leyendo archivo de confirmaciones: {e}")
        
        if not entries:
            alerts.append("No se encontraron confirmaciones válidas en el archivo")
            return [], alerts
        
        # Verificar que existe el Excel
        if not os.path.exists(self.excel_path):
            alerts.append("No existe archivo de pagos para confirmar")
            return [], alerts
        
        try:
            # Leer hoja de Pagos con dtype=str para preservar ceros a la izquierda
            df_pagos = pd.read_excel(
                self.excel_path, 
                sheet_name='Pagos', 
                engine='openpyxl',
                dtype={'ID': str, 'Ciclo': str, 'Depósito': str}
            )
            
            # Normalizar Depósito (ya es string, solo asegurar formato)
            if 'Depósito' in df_pagos.columns:
                df_pagos['Depósito'] = df_pagos['Depósito'].astype(str).str.replace('.0', '', regex=False).str.replace('nan', '').str.replace('None', '')
                # Asegurar formato completo de 9 dígitos
                def fix_deposito(val):
                    if pd.isna(val) or val == '' or val == 'nan' or val == 'None':
                        return None
                    val_str = str(val).strip()
                    if val_str.replace('.', '').isdigit():
                        val_str = val_str.split('.')[0]
                        if len(val_str) < 9:
                            val_str = val_str.zfill(9)
                    return val_str
                df_pagos['Depósito'] = df_pagos['Depósito'].apply(fix_deposito)
            
            for conf_entry in entries:
                match_found = False
                conf_tipo = conf_entry.get('Tipo', 'Gpo')  # Por defecto Gpo si no viene
                logging.info(f"Buscando confirmación: Tipo={conf_tipo}, ID={conf_entry['ID']}, Grupo={conf_entry['Grupo']}, "
                           f"Pago={conf_entry['Pago']}, Ahorro={conf_entry['Ahorro']}")
                
                # Buscar coincidencia en df_pagos con Tipo + ID + Grupo + Pago + Ahorro
                for idx in df_pagos.index:
                    row = df_pagos.iloc[idx]
                    
                    # Comparar Tipo
                    excel_tipo = str(row.get('Tipo', 'Gpo')).strip() if pd.notna(row.get('Tipo')) else 'Gpo'
                    if excel_tipo != conf_tipo:
                        continue
                    
                    # Convertir ID a string y rellenar con ceros para comparación
                    excel_id = str(row['ID']).replace('.0', '').zfill(6) if pd.notna(row['ID']) else ''
                    conf_id = str(conf_entry['ID']).strip().zfill(6)
                    
                    # Comparar ID
                    if excel_id != conf_id:
                        continue
                    
                    # Comparar Grupo (case-insensitive)
                    excel_grupo = str(row['Grupo']).strip().upper() if pd.notna(row['Grupo']) else ''
                    conf_grupo = str(conf_entry['Grupo']).strip().upper()
                    if excel_grupo != conf_grupo:
                        continue
                    
                    # Comparar Pago con tolerancia 0.01
                    excel_pago = float(row['Pago']) if pd.notna(row['Pago']) else 0.0
                    conf_pago = float(conf_entry['Pago'])
                    if abs(excel_pago - conf_pago) > 0.01:
                        logging.warning(f"Discrepancia en Pago para ID {conf_id}: "
                                      f"Excel={excel_pago} vs Confirmación={conf_pago}")
                        continue
                    
                    # Comparar Ahorro con tolerancia 0.01
                    excel_ahorro = float(row['Ahorro']) if pd.notna(row['Ahorro']) else 0.0
                    conf_ahorro = float(conf_entry['Ahorro'])
                    if abs(excel_ahorro - conf_ahorro) > 0.01:
                        logging.warning(f"Discrepancia en Ahorro para ID {conf_id}: "
                                      f"Excel={excel_ahorro} vs Confirmación={conf_ahorro}")
                    
                    # Match completo encontrado
                    logging.info(f"MATCH ENCONTRADO: Tipo={excel_tipo}, ID={conf_id}, Grupo={conf_grupo}")
                    match_found = True
                    
                    # Actualizar a "Sí" en columna Confirmado
                    df_pagos.at[idx, 'Confirmado'] = 'Sí'
                    
                    # Actualizar Ahorro si difiere
                    if abs(excel_ahorro - conf_ahorro) > 0.01:
                        df_pagos.at[idx, 'Ahorro'] = conf_ahorro
                        df_pagos.at[idx, 'Total'] = excel_pago + conf_ahorro
                    
                    # Copiar registro completo para hoja de confirmados
                    confirmed_entry = df_pagos.iloc[idx].to_dict()
                    confirmed_entries.append(confirmed_entry)
                    
                    break
                
                if not match_found:
                    alerts.append(
                        f"No se encontró: ID {conf_entry['ID']}, Grupo {conf_entry['Grupo']}, "
                        f"Pago {conf_entry['Pago']}, Ahorro {conf_entry['Ahorro']}"
                    )
            
            # Asegurar que Depósito sea string antes de guardar
            if 'Depósito' in df_pagos.columns:
                df_pagos['Depósito'] = df_pagos['Depósito'].astype(str).str.replace('.0', '', regex=False)
            
            # Guardar cambios en hoja Pagos
            with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_pagos.to_excel(writer, sheet_name='Pagos', index=False)
            
            # Configurar formato de Depósito como texto en Excel
            wb = openpyxl.load_workbook(self.excel_path)
            if 'Pagos' in wb.sheetnames:
                ws = wb['Pagos']
                for cell in ws[1]:  # Primera fila (encabezados)
                    if cell.value == 'Depósito':
                        col_letter = cell.column_letter
                        # Formatear todas las celdas de la columna Depósito como texto
                        for row in range(2, ws.max_row + 1):
                            cell_ref = ws[f'{col_letter}{row}']
                            cell_ref.number_format = '@'  # @ = texto
                            # Asegurar que el valor se guarde como string
                            if cell_ref.value is not None:
                                dep_value = str(cell_ref.value)
                                # Preservar formato completo (9 dígitos)
                                if dep_value.replace('.', '').isdigit():
                                    dep_value = dep_value.split('.')[0]
                                    if len(dep_value) < 9:
                                        dep_value = dep_value.zfill(9)
                                    cell_ref.value = dep_value
                                else:
                                    cell_ref.value = dep_value
            wb.save(self.excel_path)
            wb.close()
            
            # Actualizar hoja Pagos Confirmados
            if confirmed_entries:
                df_confirmed = pd.DataFrame(confirmed_entries)
                
                # Intentar leer confirmados existentes
                try:
                    df_existing_confirmed = pd.read_excel(
                        self.excel_path, sheet_name='Pagos Confirmados', engine='openpyxl'
                    )
                    # Combinar con los nuevos
                    df_confirmed = pd.concat([df_existing_confirmed, df_confirmed])
                except:
                    pass
                
                # Guardar hoja de confirmados
                with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_confirmed.to_excel(writer, sheet_name='Pagos Confirmados', index=False)
                
                logging.info(f"Confirmados {len(confirmed_entries)} pagos")
            
        except Exception as e:
            import traceback
            logging.error(f"Error procesando confirmaciones: {e}")
            logging.error(traceback.format_exc())
            alerts.append(f"Error procesando confirmaciones: {str(e)}")
        
        return confirmed_entries, alerts
    
    def clear_all_data(self) -> bool:
        """
        Limpia todos los registros del sistema
        Elimina Excel, limpia config y log
        Retorna True si se limpió exitosamente
        """
        errors = []
        
        # Eliminar archivo Excel con retries
        if os.path.exists(self.excel_path):
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    os.remove(self.excel_path)
                    logging.info(f"Eliminado archivo {self.excel_path}")
                    break
                except PermissionError as pe:
                    if attempt < max_retries - 1:
                        logging.warning(f"Intento {attempt + 1} de {max_retries}: Permiso denegado. Esperando...")
                        time.sleep(1)
                    else:
                        errors.append(f"NO se pudo eliminar {self.excel_path} tras {max_retries} intentos (archivo abierto en Excel)")
                except Exception as e:
                    errors.append(f"Error eliminando Excel: {e}")
                    break
        
        # Limpiar config.json
        try:
            self.config = {
                "horarios": {
                    "matutino": "< 13:00",
                    "vespertino": ">= 13:00",
                    "archivo_procesado": None,
                    "corte_actual": None
                },
                "mapeo_id_grupos": {}
            }
            self.save_config()
            logging.info("Config.json limpiado")
        except Exception as e:
            errors.append(f"Error limpiando config: {e}")
        
        # Limpiar log.txt
        try:
            if os.path.exists('log.txt'):
                os.remove('log.txt')
                logging.info("Log.txt eliminado")
        except Exception as e:
            errors.append(f"Error eliminando log: {e}")
        
        # Configurar logging de nuevo
        self.setup_logging()
        
        if errors:
            for error in errors:
                logging.warning(error)
            return False
        else:
            logging.info("Todos los datos fueron limpiados exitosamente")
            return True


def main():
    """Función principal para probar el script"""
    print("Sistema de Gestión de Pagos desde WhatsApp")
    print("=" * 50)
    
    manager = PaymentManager()
    
    # ============================================
    # OPCION PARA LIMPIAR REGISTROS (DESCOMENTAR SI SE NECESITA)
    # ============================================
    print("\nLimpiando todos los registros...")
    if manager.clear_all_data():
        print("[OK] Registros limpiados exitosamente")
    else:
        print("[ADVERTENCIA] Algunos archivos no pudieron ser eliminados")
        print("(Por ejemplo, Pagos.xlsx podría estar abierto en Excel)")
        print("El sistema continuará igualmente...")
    # ============================================
    
    # Procesar archivo de ejemplo
    filepath = "ejemplos/_chat.txt"
    
    if not os.path.exists(filepath):
        print(f"Error: No se encuentra el archivo {filepath}")
        return
    
    print(f"Procesando {filepath}...")
    entries, errors, duplicates = manager.process_file(filepath)
    
    print(f"\nResultados:")
    print(f"  Entradas extraídas: {len(entries)}")
    print(f"  Errores: {errors}")
    print(f"  Duplicados: {duplicates}")
    
    if entries:
        print(f"\nPrimeros 5 registros:")
        for i, entry in enumerate(entries[:5], 1):
            print(f"\n{i}. ID: {entry['ID']}")
            print(f"   Grupo: {entry['Grupo']}")
            print(f"   Fecha: {entry['Fecha']} {entry['Hora']}")
            print(f"   Pago: ${entry['Pago']}, Ahorro: ${entry['Ahorro']}, Total: ${entry['Total']}")
            print(f"   Sucursal: {entry['Sucursal']}")
            if entry['Número de Pago']:
                print(f"   Número de pago: {entry['Número de Pago']}")
        
        # Guardar en Excel
        print(f"\nGuardando en Excel...")
        added = manager.add_to_excel(entries)
        print(f"Guardados {added} registros en {manager.excel_path}")
        print(f"Total de entradas extraídas: {len(entries)}")
        print(f"Puedes abrir el archivo Excel para ver los resultados")
    else:
        print("No se encontraron pagos válidos")


if __name__ == "__main__":
    main()

